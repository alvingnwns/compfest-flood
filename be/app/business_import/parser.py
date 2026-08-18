from __future__ import annotations

from io import BytesIO
from typing import Any
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.business_import.schemas import ImportValidationIssue

REQUIRED_SHEETS = ("Products", "Orders", "Inventory", "Materials", "BOM")
ALLOWED_SHEETS = {*REQUIRED_SHEETS, "Instructions"}
EXPECTED_COLUMNS = {
    "Products": ("productId", "productName", "sellingPrice"),
    "Orders": ("orderId", "storeId", "productId", "quantity", "priority", "deadlineMinutes"),
    "Inventory": ("warehouseId", "productId", "availableQuantity"),
    "Materials": ("materialId", "materialName", "supplierId", "availableQuantity"),
    "BOM": ("productId", "materialId", "quantityRequired"),
}
OPTIONAL_COLUMNS = {"Products": {"unit", "category"}}
MAX_UNCOMPRESSED_SIZE = 25 * 1024 * 1024


def parse_xlsx(contents: bytes) -> tuple[dict[str, list[tuple[int, dict[str, Any]]]], list[ImportValidationIssue]]:
    issues: list[ImportValidationIssue] = []
    try:
        with ZipFile(BytesIO(contents)) as archive:
            entries = archive.infolist()
            if sum(entry.file_size for entry in entries) > MAX_UNCOMPRESSED_SIZE:
                return {}, [
                    ImportValidationIssue(
                        sheet="Workbook",
                        code="WORKBOOK_EXPANSION_LIMIT",
                        message="The workbook expands beyond the 25 MB safety limit.",
                    )
                ]
            names = [entry.filename.replace("\\", "/") for entry in entries]
            if any(name.startswith("/") or ".." in name.split("/") for name in names):
                return {}, [
                    ImportValidationIssue(
                        sheet="Workbook",
                        code="UNSAFE_ARCHIVE_PATH",
                        message="The workbook contains an unsafe internal path.",
                    )
                ]
            if any(name.casefold().endswith("vbaproject.bin") for name in names):
                return {}, [
                    ImportValidationIssue(
                        sheet="Workbook",
                        code="MACRO_NOT_ALLOWED",
                        message="Macro-enabled workbooks are not accepted.",
                    )
                ]
    except BadZipFile:
        return {}, [
            ImportValidationIssue(
                sheet="Workbook",
                code="INVALID_WORKBOOK",
                message="The file is not a readable .xlsx workbook.",
            )
        ]
    try:
        workbook = load_workbook(BytesIO(contents), read_only=True, data_only=False, keep_links=False)
    except (InvalidFileException, BadZipFile, OSError, ValueError, KeyError, EOFError):
        return {}, [
            ImportValidationIssue(
                sheet="Workbook",
                code="INVALID_WORKBOOK",
                message="The file is not a readable .xlsx workbook.",
            )
        ]

    for name in workbook.sheetnames:
        if name not in ALLOWED_SHEETS:
            issues.append(
                ImportValidationIssue(
                    sheet=name,
                    code="UNEXPECTED_SHEET",
                    message=f"Sheet '{name}' is not supported. Use the ResiliChain template.",
                )
            )
    for name in REQUIRED_SHEETS:
        if name not in workbook.sheetnames:
            issues.append(
                ImportValidationIssue(
                    sheet=name,
                    code="MISSING_SHEET",
                    message=f"Required sheet '{name}' is missing.",
                )
            )
    if issues:
        workbook.close()
        return {}, issues

    parsed: dict[str, list[tuple[int, dict[str, Any]]]] = {}
    for sheet_name in REQUIRED_SHEETS:
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows()
        header_cells = next(rows, ())
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in header_cells]
        required = EXPECTED_COLUMNS[sheet_name]
        missing = [column for column in required if column not in headers]
        allowed = set(required) | OPTIONAL_COLUMNS.get(sheet_name, set())
        unknown = [column for column in headers if column and column not in allowed]
        for column in missing:
            issues.append(
                ImportValidationIssue(
                    sheet=sheet_name,
                    row=1,
                    column=column,
                    code="MISSING_COLUMN",
                    message=f"{sheet_name} is missing required column '{column}'.",
                )
            )
        for column in unknown:
            issues.append(
                ImportValidationIssue(
                    sheet=sheet_name,
                    row=1,
                    column=column,
                    code="UNKNOWN_COLUMN",
                    message=f"{sheet_name} contains unsupported column '{column}'.",
                )
            )
        if missing or unknown:
            parsed[sheet_name] = []
            continue

        sheet_rows: list[tuple[int, dict[str, Any]]] = []
        for row_number, cells in enumerate(rows, start=2):
            values = [cell.value for cell in cells[: len(headers)]]
            if not any(value is not None and str(value).strip() for value in values):
                continue
            for index, cell in enumerate(cells[: len(headers)]):
                if cell.data_type == "f":
                    issues.append(
                        ImportValidationIssue(
                            sheet=sheet_name,
                            row=row_number,
                            column=headers[index] or None,
                            code="FORMULA_NOT_ALLOWED",
                            message=(f"{sheet_name} row {row_number} contains a formula; enter a plain value instead."),
                        )
                    )
            sheet_rows.append((row_number, dict(zip(headers, values, strict=False))))
        parsed[sheet_name] = sheet_rows
    workbook.close()
    return parsed, issues
