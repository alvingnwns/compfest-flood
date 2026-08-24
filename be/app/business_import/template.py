from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

from app.repositories.scenario_repository import get_historical_jakarta

HEADER_FILL = PatternFill("solid", fgColor="0F766E")
HEADER_FONT = Font(color="FFFFFF", bold=True)
EXAMPLE_FILL = PatternFill("solid", fgColor="ECFDF5")
THIN_BOTTOM = Border(bottom=Side(style="thin", color="CBD5E1"))


def create_business_template() -> bytes:
    demo = get_historical_jakarta()
    workbook = Workbook()
    workbook.remove(workbook.active)
    content = {
        "Products": (
            ["productId", "productName", "sellingPrice", "unit"],
            [["P001", "Frozen Chicken", 52000, "unit"], ["P002", "Fish Fillet", 68000, "unit"]],
        ),
        "Orders": (
            ["orderId", "storeId", "productId", "quantity", "priority", "deadlineMinutes"],
            [["ORDER-001", "store-a", "P001", 100, "high", 20], ["ORDER-002", "store-c", "P002", 60, "normal", 25]],
        ),
        "Inventory": (
            ["warehouseId", "productId", "availableQuantity"],
            [["wh-west", "P001", 40], ["wh-east", "P002", 25]],
        ),
        "Materials": (
            ["materialId", "materialName", "supplierId", "availableQuantity"],
            [["M001", "Chicken Raw Material", "sup-a", 150], ["M002", "Shared Packaging", "sup-b", 100]],
        ),
        "BOM": (
            ["productId", "materialId", "quantityRequired"],
            [["P001", "M001", 1], ["P001", "M002", 0.2], ["P002", "M002", 0.5]],
        ),
    }
    for sheet_name, (headers, rows) in content.items():
        sheet = workbook.create_sheet(sheet_name)
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A2"
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        for cell in sheet[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(vertical="center")
            cell.border = THIN_BOTTOM
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                cell.fill = EXAMPLE_FILL
        for column in sheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column) + 3, 30)
            sheet.column_dimensions[column[0].column_letter].width = width
        sheet.auto_filter.ref = sheet.dimensions
    workbook["Products"]["C2"].number_format = "#,##0"
    workbook["Products"]["C3"].number_format = "#,##0"

    stores = [item.id for item in demo.facilities if item.kind == "store"]
    warehouses = [item.id for item in demo.facilities if item.kind == "warehouse"]
    suppliers = [item.id for item in demo.facilities if item.kind == "supplier"]
    workbook["Orders"].add_data_validation(_list_validation("B2:B101", stores))
    workbook["Orders"].add_data_validation(_list_validation("E2:E101", ["normal", "high", "critical"]))
    workbook["Inventory"].add_data_validation(_list_validation("A2:A101", warehouses))
    workbook["Materials"].add_data_validation(_list_validation("C2:C51", suppliers))

    instructions = workbook.create_sheet("Instructions", 0)
    instructions.sheet_view.showGridLines = False
    instructions.column_dimensions["A"].width = 28
    instructions.column_dimensions["B"].width = 90
    instructions.append(
        [
            "ARUNA Business Data",
            "Replace the green example rows, keep the headers unchanged, and upload as .xlsx.",
        ]
    )
    instructions.append(
        [
            "Network scope",
            "Custom operational data runs on the current ARUNA Jakarta demo logistics network.",
        ]
    )
    instructions.append(["Priority", "normal, high, critical"])
    instructions.append(["Deadline", "Relative minutes used by the current optimizer."])
    instructions.append(["Valid storeId", ", ".join(stores)])
    instructions.append(["Valid warehouseId", ", ".join(warehouses)])
    instructions.append(["Valid supplierId", ", ".join(suppliers)])
    instructions.append(["Currency", "sellingPrice must be a numeric IDR amount per unit, for example 52000."])
    instructions.append(["Limits", "Products 20; Orders 100; Inventory 100; Materials 50; BOM 200; file size 5 MB."])
    instructions.append(["Privacy", "Raw workbook bytes are parsed in memory and are not sent to Gemini or Qwen."])
    for cell in instructions[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row in instructions.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    instructions.freeze_panes = "A2"

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _list_validation(cell_range: str, values: list[str]) -> DataValidation:
    validation = DataValidation(type="list", formula1=f'"{",".join(values)}"', allow_blank=False)
    validation.error = "Choose one of the supported Jakarta demo values."
    validation.errorTitle = "Invalid value"
    validation.add(cell_range)
    return validation
