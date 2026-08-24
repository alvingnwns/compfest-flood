from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.business_import.mapper import apply_snapshot_to_scenario
from app.business_import.service import XLSX_MIME, get_business_snapshot, import_business_workbook
from app.business_import.template import create_business_template
from app.errors import ApiError
from app.repositories.scenario_repository import get_historical_jakarta
from app.repositories.simulation_repository import simulation_repository
from app.schemas.recovery import OrderOutcome, RecoveryResult
from app.services.kpi_service import calculate_kpi
from app.services.recovery_service import RecoveryConstraints, _solve_plan


def _edited_template(mutator) -> bytes:
    workbook = load_workbook(BytesIO(create_business_template()))
    mutator(workbook)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _import_bytes(contents: bytes):
    response = import_business_workbook("business.xlsx", XLSX_MIME, contents)
    return get_business_snapshot(response.business_snapshot_id)


def _scenario_from_snapshot(snapshot):
    return apply_snapshot_to_scenario(
        get_historical_jakarta(),
        products=snapshot.products,
        orders=snapshot.orders,
        inventory=snapshot.inventory,
        materials=snapshot.materials,
    )


def test_template_download_and_valid_import_preview(client: TestClient) -> None:
    template = client.get("/api/business-data/template")
    assert template.status_code == 200
    assert "ARUNA_Business_Data_Template.xlsx" in template.headers["content-disposition"]
    workbook = load_workbook(BytesIO(template.content), data_only=False)
    assert workbook.sheetnames == ["Instructions", "Products", "Orders", "Inventory", "Materials", "BOM"]
    assert not any(cell.data_type == "f" for sheet in workbook.worksheets for row in sheet.iter_rows() for cell in row)

    imported = client.post(
        "/api/business-data/import",
        files={"file": ("business.xlsx", template.content, XLSX_MIME)},
    )
    assert imported.status_code == 201
    payload = imported.json()
    assert payload["valid"] is True
    assert payload["businessDataSource"] == "custom"
    assert payload["summary"] == {
        "productsLoaded": 2,
        "ordersLoaded": 2,
        "inventoryRows": 2,
        "materialsLoaded": 2,
        "bomRelationships": 3,
        "totalOrderValue": 9_280_000,
        "currency": "IDR",
    }
    assert client.get(f"/api/business-data/{payload['businessSnapshotId']}").status_code == 200


def test_custom_snapshot_runs_through_existing_simulation(client: TestClient) -> None:
    imported = client.post(
        "/api/business-data/import",
        files={"file": ("business.xlsx", create_business_template(), XLSX_MIME)},
    ).json()
    created = client.post(
        "/api/simulations",
        json={
            "scenarioId": "scenario-jakarta-20250304",
            "businessSnapshotId": imported["businessSnapshotId"],
        },
    )
    assert created.status_code == 201
    assert created.json()["businessDataSource"] == "custom"
    assert created.json()["businessSnapshotId"] == imported["businessSnapshotId"]
    effective = simulation_repository.get_effective_scenario(created.json()["id"])
    assert {product.id for product in effective.products} == {"P001", "P002"}
    assert {order.id for order in effective.orders} == {"ORDER-001", "ORDER-002"}
    simulation_id = created.json()["id"]
    recovery = client.post(f"/api/simulations/{simulation_id}/recovery", json={})
    assert recovery.status_code == 201
    assert recovery.json()["status"] in {"ready", "partial"}
    commerce = {
        action["orderId"]: {
            "product": action["requestedProductName"],
            "requested": action["requestedQuantity"],
            "allocated": sum(item["quantity"] for item in action["allocations"]),
            "priority": action["priority"],
        }
        for action in recovery.json()["commerceActions"]
    }
    assert commerce == {
        "ORDER-001": {"product": "Frozen Chicken", "requested": 100, "allocated": 100, "priority": "high"},
        "ORDER-002": {"product": "Fish Fillet", "requested": 60, "allocated": 60, "priority": "normal"},
    }
    unchanged = next(
        action
        for action in recovery.json()["manufacturingActions"]
        if action["baselineQuantity"] == action["recoveryQuantity"]
    )
    assert unchanged["what"].startswith("Pertahankan produksi")
    assert all(word not in unchanged["what"] for word in ("Naikkan", "Kurangi", "Sesuaikan"))
    impact = client.get(f"/api/simulations/{simulation_id}/impact")
    assert impact.status_code == 200
    assert impact.json()["simulationId"] == simulation_id
    assert impact.json()["businessDataSource"] == "custom"
    assert impact.json()["recoveryStatus"] == recovery.json()["status"]
    assert len(impact.json()["metrics"]) == 5


def test_structured_validation_errors_and_demo_safety(client: TestClient) -> None:
    def break_workbook(workbook):
        workbook.remove(workbook["BOM"])
        workbook["Orders"]["C2"] = "P999"

    response = client.post(
        "/api/business-data/import",
        files={"file": ("business.xlsx", _edited_template(break_workbook), XLSX_MIME)},
    )
    assert response.status_code == 422
    payload = response.json()
    assert payload["code"] == "BUSINESS_DATA_VALIDATION_FAILED"
    assert payload["details"]["valid"] is False
    assert any(error["code"] == "MISSING_SHEET" and error["sheet"] == "BOM" for error in payload["details"]["errors"])
    demo = client.post("/api/simulations", json={"scenarioId": "scenario-jakarta-20250304"})
    assert demo.status_code == 201
    assert demo.json()["businessDataSource"] == "demo"


def test_parser_and_validation_error_matrix() -> None:
    def missing_column(workbook):
        workbook["Products"].delete_cols(3)

    def duplicate_product(workbook):
        workbook["Products"].append(["P001", "Duplicate", 1, "unit"])

    def unknown_product(workbook):
        workbook["Orders"]["C2"] = "P999"

    def unknown_store(workbook):
        workbook["Orders"]["B2"] = "store-unknown"

    def unknown_warehouse(workbook):
        workbook["Inventory"]["A2"] = "wh-unknown"

    def negative_quantity(workbook):
        workbook["Orders"]["D2"] = -1

    def non_numeric_price(workbook):
        workbook["Products"]["C2"] = "Rp50.000"

    def invalid_priority(workbook):
        workbook["Orders"]["E2"] = "urgent"

    def invalid_deadline(workbook):
        workbook["Orders"]["F2"] = -1

    def broken_bom(workbook):
        workbook["BOM"]["B2"] = "M999"

    def unused_material(workbook):
        workbook["Materials"].append(["M999", "Unused", "sup-a", 10])

    def empty_products(workbook):
        workbook["Products"].delete_rows(2, workbook["Products"].max_row)

    def oversized_orders(workbook):
        sheet = workbook["Orders"]
        for index in range(99):
            sheet.append([f"EXTRA-{index}", "store-a", "P001", 1, "normal", 20])

    def formula_price(workbook):
        workbook["Products"]["C2"] = "=50000"

    cases = [
        (missing_column, "MISSING_COLUMN"),
        (duplicate_product, "DUPLICATE_ID"),
        (unknown_product, "UNKNOWN_PRODUCT"),
        (unknown_store, "UNKNOWN_STORE"),
        (unknown_warehouse, "UNKNOWN_WAREHOUSE"),
        (negative_quantity, "INVALID_VALUE"),
        (non_numeric_price, "INVALID_VALUE"),
        (invalid_priority, "INVALID_VALUE"),
        (invalid_deadline, "INVALID_VALUE"),
        (broken_bom, "UNKNOWN_MATERIAL"),
        (unused_material, "UNUSED_MATERIAL"),
        (empty_products, "EMPTY_SHEET"),
        (oversized_orders, "ROW_LIMIT_EXCEEDED"),
        (formula_price, "FORMULA_NOT_ALLOWED"),
    ]
    for mutator, expected_code in cases:
        with pytest.raises(ApiError) as raised:
            import_business_workbook("business.xlsx", XLSX_MIME, _edited_template(mutator))
        errors = raised.value.details["errors"]
        assert expected_code in {error["code"] for error in errors}

    with pytest.raises(ApiError) as macro_error:
        import_business_workbook("business.xlsm", XLSX_MIME, create_business_template())
    assert macro_error.value.code == "UNSUPPORTED_FILE_TYPE"


def test_unknown_snapshot_requests_reupload(client: TestClient) -> None:
    response = client.post(
        "/api/simulations",
        json={
            "scenarioId": "scenario-jakarta-20250304",
            "businessSnapshotId": "business-expired",
        },
    )
    assert response.status_code == 404
    assert response.json()["code"] == "BUSINESS_SNAPSHOT_NOT_FOUND"


def test_uploaded_price_changes_sales_exposure_exactly() -> None:
    snapshot = _import_bytes(create_business_template())
    scenario = _scenario_from_snapshot(snapshot)
    outcomes = [
        OrderOutcome(
            order_id="ORDER-001",
            requested_quantity=100,
            allocated_quantity=60,
            allocated_value=3_120_000,
            deadline_minutes=20,
            delay_minutes=0,
        ),
        OrderOutcome(
            order_id="ORDER-002",
            requested_quantity=60,
            allocated_quantity=60,
            allocated_value=4_080_000,
            deadline_minutes=25,
            delay_minutes=0,
        ),
    ]
    recovery = RecoveryResult(
        id="plan-price",
        simulation_id="sim-price",
        status="ready",
        created_at=datetime.now(UTC),
        recovery_order_outcomes=outcomes,
        baseline_order_outcomes=outcomes,
    )
    exposure = next(
        metric
        for metric in calculate_kpi("sim-price", scenario, recovery).metrics
        if metric.key == "sales-exposure-risk"
    )
    assert exposure.recovery == 40 * 52_000 == 2_080_000

    doubled = scenario.model_copy(deep=True)
    doubled.products[0].unit_price = 104_000
    doubled_exposure = next(
        metric
        for metric in calculate_kpi("sim-price", doubled, recovery).metrics
        if metric.key == "sales-exposure-risk"
    )
    assert doubled_exposure.recovery == 2 * exposure.recovery


def test_uploaded_inventory_materials_and_priority_change_solver_output(simulation_id: str) -> None:
    disruption = simulation_repository.get_disruption(simulation_id)

    def zero_inputs(workbook):
        for row in range(2, workbook["Inventory"].max_row + 1):
            workbook["Inventory"].cell(row, 3).value = 0
        for row in range(2, workbook["Materials"].max_row + 1):
            workbook["Materials"].cell(row, 4).value = 0

    high = _scenario_from_snapshot(_import_bytes(create_business_template()))
    low = _scenario_from_snapshot(_import_bytes(_edited_template(zero_inputs)))
    high_plan = _solve_plan(high, disruption, RecoveryConstraints(), baseline=False)
    low_plan = _solve_plan(low, disruption, RecoveryConstraints(), baseline=False)
    assert sum(item.allocated_quantity for item in high_plan.outcomes) > 0
    assert sum(item.allocated_quantity for item in low_plan.outcomes) == 0
    assert sum(high_plan.production.values()) > sum(low_plan.production.values())

    def priority_contention(workbook):
        orders = workbook["Orders"]
        orders.delete_rows(2, orders.max_row)
        orders.append(["ORDER-HIGH", "store-a", "P001", 400, "high", 20])
        orders.append(["ORDER-NORMAL", "store-a", "P001", 400, "normal", 20])
        for row in range(2, workbook["Inventory"].max_row + 1):
            workbook["Inventory"].cell(row, 3).value = 0
        workbook["Materials"]["D2"] = 500

    priority = _scenario_from_snapshot(_import_bytes(_edited_template(priority_contention)))
    plan = _solve_plan(priority, disruption, RecoveryConstraints(), baseline=False)
    allocated = {item.order_id: item.allocated_quantity for item in plan.outcomes}
    assert allocated["ORDER-HIGH"] > allocated["ORDER-NORMAL"]
